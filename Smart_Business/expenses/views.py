# market/expenses/views.py
from io import BytesIO
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Sum

from .models import Expense
from .forms import ExpenseForm

# Excel export
import openpyxl
from openpyxl.utils import get_column_letter

# PDF export
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


@login_required
def expense_list(request):
    """
    List + filter + pagination.
    Filter params supported:
     - q_name (search in name)
     - category (exact)
     - date_from, date_to (YYYY-MM-DD)
    """
    business = request.user.businesses.first()
    qs = Expense.objects.filter(business=business).order_by("-date", "-id")

    # --- filters ---
    q_name = request.GET.get("q_name", "").strip()
    category = request.GET.get("category", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if q_name:
        qs = qs.filter(name__icontains=q_name)
    if category:
        qs = qs.filter(category__iexact=category)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    # totals for the filtered set
    totals = qs.aggregate(total_amount=Sum("amount"))
    filtered_total = totals.get("total_amount") or 0

    # --- pagination ---
    page_size = int(request.GET.get("page_size", 15))
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # distinct categories for filter dropdown
    categories = Expense.objects.filter(business=business).order_by("category").values_list("category", flat=True).distinct()

    context = {
        "expenses": page_obj,
        "page_obj": page_obj,
        "categories": [c for c in categories if c],
        "filtered_total": filtered_total,
        "q_name": q_name,
        "category": category,
        "date_from": date_from,
        "date_to": date_to,
        "page_size": page_size,
    }
    return render(request, "expenses/expense_list.html", context)


@login_required
def expense_create(request):
    business = request.user.businesses.first()
    if request.method == "POST":
        form = ExpenseForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.business = business
            obj.save()
            messages.success(request, "Expense added successfully.")
            return redirect("expenses:expense_list")
    else:
        form = ExpenseForm()
    return render(request, "expenses/expense_form.html", {"form": form, "title": "Add Expense"})


@login_required
def expense_edit(request, pk):
    business = request.user.businesses.first()
    expense = get_object_or_404(Expense, pk=pk, business=business)
    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, "Expense updated successfully.")
            return redirect("expenses:expense_list")
    else:
        form = ExpenseForm(instance=expense)
    return render(request, "expenses/expense_form.html", {"form": form, "title": "Edit Expense"})


@login_required
def expense_delete(request, pk):
    business = request.user.businesses.first()
    expense = get_object_or_404(Expense, pk=pk, business=business)
    if request.method == "POST":
        expense.delete()
        messages.success(request, "Expense deleted successfully.")
        return redirect("expenses:expense_list")
    return render(request, "expenses/expense_delete.html", {"expense": expense})


# -----------------------
# Export helpers
# -----------------------
def _write_xlsx_response(filename, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # write headers
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # write rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # auto-width for columns
    for i, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].auto_size = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _write_pdf_response(filename, headers, rows):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Expenses Report", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    table_data = [headers] + rows
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e6e6e6")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
    ]))
    story.append(t)
    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def expense_export_xlsx(request):
    """
    Export filtered expenses to XLSX.
    Uses the same GET filter params as expense_list (q_name, category, date_from, date_to)
    """
    business = request.user.businesses.first()
    qs = Expense.objects.filter(business=business).order_by("-date", "-id")

    # apply same filters
    q_name = request.GET.get("q_name", "").strip()
    category = request.GET.get("category", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if q_name:
        qs = qs.filter(name__icontains=q_name)
    if category:
        qs = qs.filter(category__iexact=category)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    headers = ["Date", "Name", "Category", "Amount", "Notes"]
    rows = []
    for e in qs:
        rows.append([e.date.strftime("%Y-%m-%d"), e.name, e.category or "", float(e.amount), e.notes or ""])

    filename = f"expenses_{business.pk}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _write_xlsx_response(filename, headers, rows)


@login_required
def expense_export_pdf(request):
    """
    Export filtered expenses to PDF.
    """
    business = request.user.businesses.first()
    qs = Expense.objects.filter(business=business).order_by("-date", "-id")

    # apply filters (same as above)
    q_name = request.GET.get("q_name", "").strip()
    category = request.GET.get("category", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if q_name:
        qs = qs.filter(name__icontains=q_name)
    if category:
        qs = qs.filter(category__iexact=category)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    headers = ["Date", "Name", "Category", "Amount", "Notes"]
    rows = []
    for e in qs:
        rows.append([e.date.strftime("%Y-%m-%d"), e.name, e.category or "", f"{e.amount}", e.notes or ""])

    filename = f"expenses_{business.pk}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return _write_pdf_response(filename, headers, rows)
